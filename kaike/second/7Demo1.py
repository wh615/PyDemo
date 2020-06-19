# pip install openpyxl -i https://pypi.douban.com/simple

import openpyxl
# 引用openpyxl
wb = openpyxl.Workbook()
# 利用openpyxl.Workbook()函数创建新的workbook（工作薄）对象，就是创建新的空的Excel文件。
sheet = wb.active
# wb.active就是获取这个工作薄的活动表，通常就是第一个工作簿，也就是我们在上面的图片中看到的sheet1。
sheet.title = 'kaikeba'
# 可以用.title给工作表重命名。现在第一个工作表的名称就会由原来默认的“sheet1”改为"kaikeba"。
sheet['A1'] = 'kaikeba'
sheet['B1'] = '分数'
# 向单个单元格写入数据
score1 = ['math', 95]
sheet.append(score1)
# 写入整行的数据，变量类型是一个列表
wb.save('score.xlsx')
# 保存修改的Excel
wb.close()
# 关闭Excel

wb=openpyxl.load_workbook('score.xlsx')
sheet=wb['kaikeba']
print(sheet['A1'].value)

